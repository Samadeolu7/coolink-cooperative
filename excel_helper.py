import pandas as pd, os, csv, zipfile
from flask import Flask
from models import db, Person
from queries import Queries
from openpyxl import Workbook
from openpyxl.styles import Alignment
from filters import format_currency
from sqlalchemy.orm import Session

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///cooperative.db"
db.init_app(app)

query = Queries(db)


def log_report(report):
    with open("report.txt", "a", encoding="utf-8") as f:
        f.write(f"{report}\n")


# processing input functions
def create_excel(type, type_id):
    if type == "savings":
        if type_id == "None":
            return create_all_savings_excel()
        else:
            person = query.get_person(type_id)
            return create_payments_excel(person)

    elif type == "loan":
        if type_id == "None":
            return create_all_loans_excel()
        else:
            person = query.get_person(type_id)
            return create_loan_excel(person)

    elif type == "bank":
        if type_id == "None":
            return
        else:
            bank = query.get_bank(type_id)
            return create_banks_excel(bank)

    elif type == "income":
        if type_id == "None":
            return create_income_excel()
        else:
            return
    elif type == "persons":
        if type_id == "None":
            return create_persons_excel()


def process_excel(filename):
    dataframe = pd.read_excel(filename)
    rows_to_update = dataframe.iloc[10:]
    return rows_to_update


def start_up(filename):
    df = process_excel(filename)
    records = []

    for _, row in df.iterrows():
        password = query.generate_password()
        employee_id = row["COY"]
        email =f'testemail{row["S/N"]}@gmail.com'
        record = {
            "name": row["Name"],
            "employee_id": employee_id,
            "email": email,
            "password": password,
            "available_balance": row["BAL B/FWD"],
            'loan_balance': 0,  # Adjust the column name as needed
            'loan_balance_bfd': 0,  # Adjust the column name as needed
            "phone_no": row["Phone"],
            "balance_bfd": row["BAL B/FWD"],
            "company_id": 1,  # Adjust the company ID as needed
            "role_id": 4,  # Assuming a default role_id of 4 for new users
        }
        records.append(record)

        with open(
            f"credentials/{employee_id}_credentials.csv", "w", newline=""
        ) as file:
            writer = csv.writer(file)
            writer.writerow(["Username", "Password"])
            writer.writerow([f"{employee_id} or {email}", password])

    db.session.bulk_insert_mappings(Person, records)
    db.session.commit()
    zip_filename = "credentials.zip"
    with zipfile.ZipFile(zip_filename, "w") as zipf:
        for employee_id in df["COY"]:
            csv_filename = f"credentials/{employee_id}_credentials.csv"
            zipf.write(csv_filename, f"{employee_id}_credentials.csv")

    return zip_filename


def send_upload_to_savings(filename,ref_no, description, date):
    df = process_excel(filename)
    error =[]
    for index, row in df.iterrows():
        test=query.save_amount_company(
            row["COY"], row["Amount"], date, ref_no, description
        )
        if test != True:
            error.append(test)
    
    if len(error) > 0:
        return error
    else:
        return True


def send_upload_to_loan_repayment(filename,ref_no, description, date):
    df = process_excel(filename)
    for index, row in df.iterrows():
        query.repay_loan_company(
            row["COY"],
            row["Amount"],
            date,
            ref_no,
            description,
        )


# Function to generate the repayment schedule
def generate_repayment_schedule(
    person_id, loan_amount, interest_rate, start_date, end_date
):
    # Calculate the number of months between the start date and end date
    num_months = (
        (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month) + 1
    )

    # Calculate the total interest amount
    total_interest = loan_amount * (int(interest_rate) / 100)

    # Calculate the monthly repayment amount
    monthly_repayment = (loan_amount + total_interest) / num_months

    # Generate the repayment schedule as a list of dictionaries
    repayment_schedule = []
    for i in range(num_months):
        installment_number = i + 1
        due_date = start_date + pd.DateOffset(months=i)
        repayment_amount = monthly_repayment
        repayment_schedule.append(
            {
                "Installment Number": installment_number,
                "Due Date": due_date,
                "Repayment Amount": repayment_amount,
            }
        )

    return repayment_schedule


# Function to export the repayment schedule to Excel
def export_repayment_schedule_to_excel(repayment_schedule, person_id):
    # Create a DataFrame from the repayment schedule
    df_repayment_schedule = pd.DataFrame(repayment_schedule)

    # Generate a unique file name based on the person ID
    file_name = f"repayment_schedule_{person_id}.xlsx"
    file_path = os.path.join("./excel", file_name)

    # Create an Excel writer object
    writer = pd.ExcelWriter(file_path, engine="xlsxwriter")

    # Write the repayment schedule DataFrame to the Excel file
    df_repayment_schedule.to_excel(writer, sheet_name="Repayment Schedule", index=False)

    # Save the Excel file
    writer.close()

    return file_path


def save(rows_to_update):
    for index, row in rows_to_update.iterrows():
        query.save_amount(row["COY"], row["Amount"])


# processing ouputs functions


def create_individual_report(employee_id):
    person = query.get_person(employee_id)
    df_person = pd.DataFrame(
        [
            (
                person.id,
                person.is_admin,
                person.employee_id,
                person.name,
                person.email,
                person.phone_no,
                person.savings,
                person.available_balance,
                person.loan_balance,
                person.monthly_payment_amount,
                person.company_id,
            )
            for person in person
        ],
        columns=[
            "ID",
            "Is Admin",
            "Employee ID",
            "Name",
            "Email",
            "Phone No",
            "Savings",
            "Total Balance",
            "Loan Balance",
            "Monthly Payment Amount",
            "Company ID",
        ],
    )

    file_name = "cooperative_data.xlsx"
    file_path = os.path.join("./excel", file_name)

    # Create an Excel writer object
    writer = pd.ExcelWriter(file_path, engine="xlsxwriter")

    df_person.to_excel(writer, sheet_name="Persons", index=False)
    writer.close()


def create_full_report():
    # Fetch data from the database
    companies = query.get_companies()
    persons = query.get_persons()
    payments = query.get_payments()
    loans = query.get_loans()
    payments = query.get_payments()
    investments = query.get_investments()
    expenses = query.get_expenses()

    # Create dataframes from the fetched data
    # remember to add amount accumulated to the company

    df_companies = pd.DataFrame(
        [(company.id, company.name) for company in companies], columns=["ID", "Name"]
    )
    df_persons = pd.DataFrame(
        [
            (
                person.id,
                person.is_admin,
                person.employee_id,
                person.name,
                person.email,
                person.phone_no,
                person.savings,
                person.available_balance,
                person.loan_balance,
                person.monthly_payment_amount,
                person.company_id,
            )
            for person in persons
        ],
        columns=[
            "ID",
            "Is Admin",
            "Employee ID",
            "Name",
            "Email",
            "Phone No",
            "Savings",
            "Total Balance",
            "Loan Balance",
            "Monthly Payment Amount",
            "Company ID",
        ],
    )
    df_payments = pd.DataFrame(
        [
            (payment.id, payment.amount, payment.loan, payment.date, payment.person_id)
            for payment in payments
        ],
        columns=["ID", "Amount", "Loan", "Date", "Person ID"],
    )
    df_loans = pd.DataFrame(
        [
            (
                loan.id,
                loan.person_id,
                loan.amount,
                loan.interest_rate,
                loan.start_date,
                loan.end_date,
                loan.is_paid,
            )
            for loan in loans
        ],
        columns=[
            "ID",
            "Person ID",
            "Amount",
            "Interest Rate",
            "Start Date",
            "End Date",
            "Is Paid",
        ],
    )
    df_expenses = pd.DataFrame(
        [
            (expense.id, expense.description, expense.amount, expense.date)
            for expense in expenses
        ],
        columns=["ID", "Description", "Amount", "Date"],
    )
    df_investments = pd.DataFrame(
        [
            (investment.id, investment.description, investment.amount, investment.date)
            for investment in investments
        ],
        columns=["ID", "Description", "Amount", "Date"],
    )

    # Create an Excel writer object
    file_name = "cooperative_data.xlsx"
    file_path = os.path.join("./excel", file_name)

    # Create an Excel writer object
    writer = pd.ExcelWriter(file_path, engine="xlsxwriter")

    # Write dataframes to separate sheets in the Excel file
    df_companies.to_excel(writer, sheet_name="Companies", index=False)
    df_persons.to_excel(writer, sheet_name="Persons", index=False)
    df_payments.to_excel(writer, sheet_name="Payments", index=False)
    df_loans.to_excel(writer, sheet_name="Loans", index=False)
    df_expenses.to_excel(writer, sheet_name="Expenses", index=False)
    df_investments.to_excel(writer, sheet_name="Investments", index=False)

    # Save the Excel file
    writer.close()


def create_persons_excel():
    # Create a new workbook and select the active sheet

    workbook = Workbook()
    sheet = workbook.active

    # Set column widths
    column_widths = [15, 15, 30, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # Add person information
    sheet["A1"] = "Coolink Cooperative Savings Account"
    sheet["A2"] = ""
    sheet["A3"] = ""
    sheet["A4"] = ""

    # Add header row
    header = [
        "Employee ID",
        "Name",
        "Email",
        "Phone Number",
        "Savings Balance",
        "Loan  Balance",
        "Company",
    ]
    for col_num, header_value in enumerate(header, start=1):
        cell = sheet.cell(row=6, column=col_num)
        cell.value = header_value
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add payments made rows
    for person in query.get_persons():
        payment_row = [
            person.employee_id,
            person.name,
            person.email,
            person.phone_no,
            format_currency(person.available_balance),
            format_currency(person.loan_balance),
            person.company.name,
        ]
        sheet.append(payment_row)

    # Save the workbook
    file_path = "excel/persons.xlsx"
    workbook.save(file_path)

    return file_path


def create_payments_excel(person):
    # Create a new workbook and select the active sheet

    workbook = Workbook()
    sheet = workbook.active

    # Set column widths
    column_widths = [15, 15, 30, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # Add person information
    sheet["A1"] = "Coolink Cooperative Savings Account"
    sheet["A2"] = f"Name: {person.name}"
    sheet["A3"] = f"Company: {person.company.name}"
    sheet["A4"] = f"Member ID: {person.employee_id}"

    # Add header row
    header = ["Date", "Reference Number", "Description", "Amount", "Balance"]
    for col_num, header_value in enumerate(header, start=1):
        cell = sheet.cell(row=6, column=col_num)
        cell.value = header_value
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add balance B/FWD row
    balance_bfd_row = [
        "None",
        "None",
        "Balance B/FWD",
        format_currency(person.balance_bfd),
        format_currency(person.balance_bfd),
    ]
    sheet.append(balance_bfd_row)

    # Add payments made rows
    for payment in person.payments_made:
        payment_row = [
            payment.date.strftime("%Y-%m-%d"),
            payment.ref_no,
            payment.description,
            format_currency(payment.amount),
            format_currency(payment.balance),
        ]
        sheet.append(payment_row)

    # Add total balance row
    available_balance_row = [
        "None",
        "None",
        "None",
        "None",
        format_currency(person.available_balance),
    ]
    sheet.append(available_balance_row)

    # Save the workbook
    file_path = "excel/person_payments.xlsx"
    workbook.save(file_path)

    return file_path


def create_loan_excel(person):
    # Create a new workbook and select the active sheet

    workbook = Workbook()
    sheet = workbook.active

    # Set column widths
    column_widths = [15, 15, 30, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # Add person information
    sheet["A1"] = "Coolink Cooperative Loan Account"
    sheet["A2"] = f"Name: {person.name}"
    sheet["A3"] = f"Company: {person.company.name}"
    sheet["A4"] = f"Member ID: {person.employee_id}"

    # Add header row
    header = ["Date", "Reference Number", "Description", "Amount", "Balance"]
    for col_num, header_value in enumerate(header, start=1):
        cell = sheet.cell(row=6, column=col_num)
        cell.value = header_value
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add balance B/FWD row
    balance_bfd_row = [
        "None",
        "None",
        "Balance B/FWD",
        format_currency(person.loan_balance_bfd),
        format_currency(person.loan_balance_bfd),
    ]
    sheet.append(balance_bfd_row)

    # Add payments made rows
    for payment in person.loan_payments_made:
        payment_row = [
            payment.date.strftime("%Y-%m-%d"),
            payment.ref_no,
            payment.description,
            format_currency(payment.amount),
            format_currency(payment.balance),
        ]
        sheet.append(payment_row)

    # Add total balance row
    available_balance_row = [
        "None",
        "None",
        "None",
        "None",
        format_currency(person.loan_balance),
    ]
    sheet.append(available_balance_row)

    # Save the workbook
    file_path = f"excel/{person.name}_payments.xlsx"
    workbook.save(file_path)

    return file_path


def create_all_savings_excel():
    # Create a new workbook and select the active sheet

    workbook = Workbook()
    sheet = workbook.active

    # Set column widths
    column_widths = [15, 15, 30, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # Add person information
    sheet["A1"] = "Coolink Cooperative Savings Account"
    sheet["A2"] = " "

    # Add header row
    header = ["Date", "Reference Number", "Description", "Amount"]
    for col_num, header_value in enumerate(header, start=1):
        cell = sheet.cell(row=6, column=col_num)
        cell.value = header_value
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add payments made rows
    for payment in query.get_savings():
        payment_row = [
            payment.date.strftime("%Y-%m-%d"),
            payment.ref_no,
            payment.description,
            format_currency(payment.amount),
        ]
        sheet.append(payment_row)

    # Save the workbook
    file_path = "excel/savings.xlsx"
    workbook.save(file_path)

    return file_path


def create_all_loans_excel():
    # Create a new workbook and select the active sheet

    workbook = Workbook()
    sheet = workbook.active

    # Set column widths
    column_widths = [15, 15, 30, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # Add person information
    sheet["A1"] = "Coolink Cooperative Loans Account"
    sheet["A2"] = " "

    # Add header row
    header = ["Date", "Reference Number", "Description", "Amount"]
    for col_num, header_value in enumerate(header, start=1):
        cell = sheet.cell(row=6, column=col_num)
        cell.value = header_value
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add payments made rows
    for payment in query.get_loans():
        payment_row = [
            payment.id,
            payment.person.name,
            format_currency(payment.amount),
            payment.interest_rate,
            payment.person.loan_balance,
            payment.start_date,
            payment.end_date,
            payment.is_paid,
        ]
        sheet.append(payment_row)

    # Save the workbook
    file_path = "excel/loans.xlsx"
    workbook.save(file_path)

    return file_path


def create_banks_excel(bank):
    # Create a new workbook and select the active sheet

    workbook = Workbook()
    sheet = workbook.active

    # Set column widths
    column_widths = [15, 15, 30, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # Add person information
    sheet["A1"] = "Coolink Cooperative Savings Account"
    sheet["A2"] = " "

    # Add header row
    header = ["Date", "Reference Number", "Description", "Amount"]
    for col_num, header_value in enumerate(header, start=1):
        cell = sheet.cell(row=6, column=col_num)
        cell.value = header_value
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add payments made rows
    for payment in query.get_bank(bank.id).payments:
        payment_row = [
            payment.date.strftime("%Y-%m-%d"),
            payment.ref_no,
            payment.description,
            format_currency(payment.amount),
            format_currency(payment.bank_balance),
        ]
        sheet.append(payment_row)

    # Save the workbook
    file_path = "excel/savings.xlsx"
    workbook.save(file_path)

    return file_path


def create_income_excel():
    # Create a new workbook and select the active sheet

    workbook = Workbook()
    sheet = workbook.active

    # Set column widths
    column_widths = [15, 15, 30, 15, 15]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[chr(64 + i)].width = width

    # Add person information
    sheet["A1"] = "Coolink Cooperative Income Account"
    sheet["A2"] = " "

    # Add header row
    header = ["Name", "Description", "Amount"]
    for col_num, header_value in enumerate(header, start=1):
        cell = sheet.cell(row=6, column=col_num)
        cell.value = header_value
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add payments made rows
    income_row = ['Income']
    sheet.append(income_row)
    for payment in query.get_income():
        payment_row = [
            payment.name,
            payment.description,
            format_currency(payment.balance),
        ]
        sheet.append(payment_row)

    expense_row = ['Expense']
    sheet.append(expense_row)
    for payment in query.get_expenses():
        payment_row = [
            payment.name,
            payment.description,
            format_currency(payment.balance),
        ]
        sheet.append(payment_row)
    
    net_income = query.get_net_income()
    net_income_row = ['Net Income', '', format_currency(net_income)]
    sheet.append(net_income_row)

    # Save the workbook
    file_path = "excel/income.xlsx"
    workbook.save(file_path)

    return file_path

with app.app_context():
    process_excel("./Cooperative 2022 Financial.xlsx")
